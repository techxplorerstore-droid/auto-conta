"""Extracción de texto crudo desde facturas en Excel (openpyxl).

Contrato: devuelve SOLO texto normalizado (str). No estructura ni parsea a
`Factura` — eso es responsabilidad de la capa de IA (lote LLM).

Estrategia: en vez de asumir una tabla limpia (los layouts de factura suelen
mezclar encabezado, ítems y totales en una misma hoja), recorremos las filas
de la hoja activa y serializamos cada fila como "celda | celda | ...",
saltando celdas vacías. Esto es más robusto que `read_excel` con header fijo.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def _celda_a_texto(valor: object) -> str:
    """Convierte un valor de celda a texto, sin decimales superfluos en enteros."""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor)


def extraer_texto(path: Path) -> str:
    """Devuelve una representación textual fiel del contenido de un Excel.

    Cada fila no vacía se serializa como "c1 | c2 | ...", uniendo filas con
    saltos de línea.

    Raises:
        FileNotFoundError: si el archivo no existe.
        ValueError: si la hoja está vacía / sin contenido legible.
    """
    if not path.exists():
        raise FileNotFoundError(f"El archivo Excel no existe: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        lineas: list[str] = []
        for fila in ws.iter_rows(values_only=True):
            celdas = [_celda_a_texto(c).strip() for c in fila if c is not None]
            celdas = [c for c in celdas if c]
            if celdas:
                lineas.append(" | ".join(celdas))
    finally:
        wb.close()

    if not lineas:
        raise ValueError(f"No se pudo extraer contenido del Excel: {path}")

    return "\n".join(lineas).strip()
