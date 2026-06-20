"""Genera facturas colombianas SINTÉTICAS (datos 100% ficticios) en samples/.

Produce 5 PDF (fpdf2) y 2 Excel (openpyxl) con proveedores y categorías
variadas, para que el clasificador del Lote 3 tenga diversidad con qué probar.

NOTA: ningún NIT, nombre, dirección o monto corresponde a entidades reales.

Uso:
    python samples/generar_muestras.py
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import List, Tuple

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook

# Equivalente moderno de `ln=True` (saltar a la siguiente línea desde el margen).
_NL = {"new_x": XPos.LMARGIN, "new_y": YPos.NEXT}

SAMPLES_DIR = Path(__file__).resolve().parent
IVA_PORCENTAJE = Decimal("19")


@dataclass
class Item:
    descripcion: str
    cantidad: Decimal
    valor_unitario: Decimal

    @property
    def valor_total(self) -> Decimal:
        return self.cantidad * self.valor_unitario


@dataclass
class FacturaMuestra:
    archivo: str
    numero_factura: str
    proveedor: str
    nit: str
    direccion: str
    ciudad: str
    fecha: str  # ISO YYYY-MM-DD (ficticia)
    categoria: str
    items: List[Item] = field(default_factory=list)

    @property
    def subtotal(self) -> Decimal:
        return sum((it.valor_total for it in self.items), Decimal("0"))

    @property
    def iva_valor(self) -> Decimal:
        return (self.subtotal * IVA_PORCENTAJE / Decimal("100")).quantize(Decimal("1"))

    @property
    def total(self) -> Decimal:
        return self.subtotal + self.iva_valor


def cop(valor: Decimal) -> str:
    """Formatea un Decimal como pesos colombianos: $ 1.234.567."""
    entero = int(valor.quantize(Decimal("1")))
    return "$ " + f"{entero:,}".replace(",", ".")


# --- Datos ficticios -------------------------------------------------------

FACTURAS: List[FacturaMuestra] = [
    FacturaMuestra(
        archivo="factura_papeleria.pdf",
        numero_factura="FE-2026-00187",
        proveedor="Papelería El Lápiz Dorado S.A.S.",
        nit="900.451.882-3",
        direccion="Calle 45 # 12-30, Local 4",
        ciudad="Bogotá D.C.",
        fecha="2026-02-11",
        categoria="papeleria_oficina",
        items=[
            Item("Resma papel carta 75g (x500)", Decimal("10"), Decimal("18500")),
            Item("Esfero tinta negra caja x12", Decimal("4"), Decimal("22000")),
            Item("Carpeta legajadora oficio", Decimal("25"), Decimal("3200")),
            Item("Tóner compatible impresora láser", Decimal("2"), Decimal("145000")),
        ],
    ),
    FacturaMuestra(
        archivo="factura_energia.pdf",
        numero_factura="EE-887234561",
        proveedor="Empresa de Energía Andina E.S.P.",
        nit="830.112.998-1",
        direccion="Carrera 7 # 26-20, Piso 12",
        ciudad="Bogotá D.C.",
        fecha="2026-02-05",
        categoria="servicios_publicos",
        items=[
            Item("Consumo energía febrero (kWh 1.240)", Decimal("1"), Decimal("612480")),
            Item("Cargo fijo comercial", Decimal("1"), Decimal("28900")),
            Item("Alumbrado público", Decimal("1"), Decimal("19450")),
        ],
    ),
    FacturaMuestra(
        archivo="factura_arriendo.pdf",
        numero_factura="ARR-2026-014",
        proveedor="Inmobiliaria Casa Verde Ltda.",
        nit="811.305.447-6",
        direccion="Avenida El Poblado # 9-15, Of. 301",
        ciudad="Medellín",
        fecha="2026-02-01",
        categoria="arriendo",
        items=[
            Item("Canon arrendamiento oficina 302 - febrero", Decimal("1"), Decimal("3800000")),
            Item("Cuota administración", Decimal("1"), Decimal("420000")),
        ],
    ),
    FacturaMuestra(
        archivo="factura_transporte.pdf",
        numero_factura="TRP-55218",
        proveedor="Transportes Rápido Ya S.A.S.",
        nit="901.778.213-9",
        direccion="Km 4 vía Cali - Yumbo, Bodega 7",
        ciudad="Cali",
        fecha="2026-02-18",
        categoria="transporte",
        items=[
            Item("Flete mensajería urbana (x20 envíos)", Decimal("20"), Decimal("14500")),
            Item("Transporte de carga Bogotá-Cali", Decimal("1"), Decimal("980000")),
            Item("Recargo combustible", Decimal("1"), Decimal("76000")),
        ],
    ),
    FacturaMuestra(
        archivo="factura_tecnologia.pdf",
        numero_factura="SOL-2026-0421",
        proveedor="SoluTech Colombia S.A.S.",
        nit="901.204.665-2",
        direccion="Calle 100 # 19-54, Torre B Of. 802",
        ciudad="Bogotá D.C.",
        fecha="2026-02-22",
        categoria="tecnologia",
        items=[
            Item("Licencia anual software gestión (x5 usuarios)", Decimal("5"), Decimal("420000")),
            Item("Soporte técnico mensual", Decimal("1"), Decimal("350000")),
            Item("Disco SSD 1TB", Decimal("3"), Decimal("298000")),
        ],
    ),
]

FACTURAS_EXCEL: List[FacturaMuestra] = [
    FacturaMuestra(
        archivo="factura_mantenimiento.xlsx",
        numero_factura="MNT-2026-0099",
        proveedor="Aseo y Mantenimiento Brillante S.A.S.",
        nit="900.667.120-4",
        direccion="Diagonal 61 # 14-22",
        ciudad="Bogotá D.C.",
        fecha="2026-02-14",
        categoria="mantenimiento",
        items=[
            Item("Servicio de aseo general (mensual)", Decimal("1"), Decimal("1250000")),
            Item("Mantenimiento aire acondicionado", Decimal("2"), Decimal("180000")),
            Item("Insumos de limpieza", Decimal("1"), Decimal("95000")),
        ],
    ),
    FacturaMuestra(
        archivo="factura_honorarios.xlsx",
        numero_factura="HON-2026-031",
        proveedor="Consultores Asociados JM Ltda.",
        nit="860.998.331-7",
        direccion="Carrera 15 # 88-64, Of. 502",
        ciudad="Bogotá D.C.",
        fecha="2026-02-20",
        categoria="honorarios",
        items=[
            Item("Asesoría contable y tributaria - febrero", Decimal("1"), Decimal("2400000")),
            Item("Preparación declaración de renta", Decimal("1"), Decimal("850000")),
        ],
    ),
]


# --- Generadores -----------------------------------------------------------

def generar_pdf(f: FacturaMuestra) -> Path:
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Encabezado: proveedor
    pdf.set_font("helvetica", "B", 15)
    pdf.cell(0, 9, f.proveedor, **_NL)
    pdf.set_font("helvetica", "", 10)
    pdf.cell(0, 5, f"NIT: {f.nit}", **_NL)
    pdf.cell(0, 5, f"{f.direccion} - {f.ciudad}", **_NL)
    pdf.ln(3)

    # Datos de la factura
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 7, "FACTURA ELECTRONICA DE VENTA", **_NL)
    pdf.set_font("helvetica", "", 10)
    pdf.cell(0, 5, f"No. {f.numero_factura}", **_NL)
    pdf.cell(0, 5, f"Fecha de emision: {f.fecha}", **_NL)
    pdf.ln(4)

    # Tabla de ítems
    pdf.set_font("helvetica", "B", 9)
    pdf.cell(95, 7, "Descripcion", border=1)
    pdf.cell(18, 7, "Cant.", border=1, align="R")
    pdf.cell(36, 7, "V. Unitario", border=1, align="R")
    pdf.cell(36, 7, "V. Total", border=1, align="R", **_NL)

    pdf.set_font("helvetica", "", 9)
    for it in f.items:
        pdf.cell(95, 6, it.descripcion, border=1)
        pdf.cell(18, 6, str(it.cantidad), border=1, align="R")
        pdf.cell(36, 6, cop(it.valor_unitario), border=1, align="R")
        pdf.cell(36, 6, cop(it.valor_total), border=1, align="R", **_NL)

    pdf.ln(3)

    # Totales
    pdf.set_font("helvetica", "", 10)
    pdf.cell(149, 6, "Subtotal", align="R")
    pdf.cell(36, 6, cop(f.subtotal), align="R", **_NL)
    pdf.cell(149, 6, f"IVA ({IVA_PORCENTAJE}%)", align="R")
    pdf.cell(36, 6, cop(f.iva_valor), align="R", **_NL)
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(149, 7, "TOTAL A PAGAR", align="R")
    pdf.cell(36, 7, cop(f.total), align="R", **_NL)

    pdf.ln(6)
    pdf.set_font("helvetica", "I", 8)
    pdf.cell(0, 4, "Documento sintetico generado para pruebas. Datos ficticios.", **_NL)

    ruta = SAMPLES_DIR / f.archivo
    pdf.output(str(ruta))
    return ruta


def generar_excel(f: FacturaMuestra) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Factura"

    # Encabezado
    ws.append([f.proveedor])
    ws.append([f"NIT: {f.nit}"])
    ws.append([f"{f.direccion} - {f.ciudad}"])
    ws.append([])
    ws.append(["Factura No.", f.numero_factura])
    ws.append(["Fecha de emision", f.fecha])
    ws.append(["Categoria", f.categoria])
    ws.append([])

    # Tabla de ítems
    ws.append(["Descripcion", "Cantidad", "Valor unitario", "Valor total"])
    for it in f.items:
        ws.append([
            it.descripcion,
            float(it.cantidad),
            float(it.valor_unitario),
            float(it.valor_total),
        ])
    ws.append([])
    ws.append(["", "", "Subtotal", float(f.subtotal)])
    ws.append(["", "", f"IVA ({IVA_PORCENTAJE}%)", float(f.iva_valor)])
    ws.append(["", "", "TOTAL", float(f.total)])

    # Ancho de columnas para legibilidad
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    ruta = SAMPLES_DIR / f.archivo
    wb.save(str(ruta))
    return ruta


def main() -> None:
    generados: List[Tuple[str, Path]] = []
    for f in FACTURAS:
        generados.append(("PDF", generar_pdf(f)))
    for f in FACTURAS_EXCEL:
        generados.append(("XLSX", generar_excel(f)))

    print(f"Generadas {len(generados)} facturas sinteticas en {SAMPLES_DIR}:")
    for tipo, ruta in generados:
        print(f"  [{tipo}] {ruta.name}")


if __name__ == "__main__":
    main()
